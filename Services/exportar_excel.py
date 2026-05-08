import io
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from plot.graficas import (
    graficar_metodo_cerrado,
    graficar_newton,
    graficar_secante,
    graficar_punto_fijo,
    graficar_taylor,
    graficar_muller
)


# 🔹 PROCESADOR UNIVERSAL
def procesar_fila_compleja(fila):

    nueva_fila = []

    for celda in fila:

        # Tuplas
        if isinstance(celda, tuple):
            nueva_fila.append(f"[{celda[0]:.4f}, {celda[1]:.4f}]")

        # Listas especiales
        elif (
            isinstance(celda, list)
            and len(celda) > 0
            and isinstance(celda[-1], dict)
        ):

            ult = celda[-1]

            res = (
                ult.get('raiz')
                or ult.get('Ci+1')
                or ult.get('x_nuevo')
                or 0
            )

            err = (
                ult.get('Error%')
                or ult.get('error')
                or 0
            )

            # Resultado complejo
            if isinstance(res, complex):

                if abs(res.imag) < 1e-10:
                    nueva_fila.append(round(float(res.real), 6))

                else:
                    signo = "+" if res.imag >= 0 else "-"

                    nueva_fila.append(
                        f"{res.real:.6f} "
                        f"{signo} "
                        f"{abs(res.imag):.6f}i"
                    )

            else:
                nueva_fila.append(round(float(res), 6))

            nueva_fila.append(f"{err:.4e}")

        # Complejos normales
        elif isinstance(celda, complex):

            # Si la parte imaginaria es prácticamente cero
            if abs(celda.imag) < 1e-10:
                nueva_fila.append(round(float(celda.real), 8))

            # Complejo real
            else:
                signo = "+" if celda.imag >= 0 else "-"

                nueva_fila.append(
                    f"{celda.real:.8f} "
                    f"{signo} "
                    f"{abs(celda.imag):.8f}i"
                )

        # Float normales
        elif isinstance(celda, (float, np.float64, np.float32)):
            nueva_fila.append(round(float(celda), 8))

        # Enteros NumPy
        elif isinstance(celda, (np.int32, np.int64)):
            nueva_fila.append(int(celda))

        # Todo lo demás
        else:
            nueva_fila.append(celda)

    return nueva_fila


# 🔹 EXPORTADOR GENÉRICO
def exportar_excel_generico(
    df,
    f_num=None,
    metodo_nombre="Reporte",
    iteraciones=None,
    extra_params=None
):

    output = io.BytesIO()

    wb = Workbook()
    ws = wb.active
    ws.title = metodo_nombre[:30]

    # Convertir listas a DataFrame
    if isinstance(df, list):
        df = pd.DataFrame(df)

    # Estilos
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    ws.append(df.columns.tolist())

    # Filas
    for _, fila in df.iterrows():
        ws.append(procesar_fila_compleja(fila))

    # Estilos de tabla
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):

        for cell in row:

            cell.border = thin_border

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font

    # Autoajuste de columnas
    for column in ws.columns:

        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column
        )

        ws.column_dimensions[
            column[0].column_letter
        ].width = max_length + 4

    # 🔹 GRÁFICAS
    try:

        fig = None

        # Métodos cerrados
        if metodo_nombre in ["Biseccion", "ReglaFalsa"] and iteraciones:

            fig = graficar_metodo_cerrado(
                f_num,
                iteraciones,
                metodo_nombre
            )

        # Newton-Raphson
        elif metodo_nombre == "NewtonRaphson" and iteraciones:

            fig = graficar_newton(
                f_num,
                iteraciones
            )

        # Secante
        elif metodo_nombre == "Secante" and iteraciones:

            fig = graficar_secante(
                f_num,
                iteraciones
            )

        # Punto fijo
        elif metodo_nombre == "PuntoFijo" and extra_params:

            fig = graficar_punto_fijo(
                extra_params["gs"],
                iteraciones,
                extra_params["x_min"],
                extra_params["x_max"]
            )
            
        # Taylor
        elif metodo_nombre == "Taylor" and extra_params:

            fig = graficar_taylor(
                f_num,
                extra_params["poly_func"],
                extra_params["x_eval"],
                extra_params["a"]
            )

        # Muller
        elif metodo_nombre == "Muller" and iteraciones:

            fig = graficar_muller(
                f_num,
                iteraciones
            )

        # Insertar gráfica
        if fig:

            img_bytes = io.BytesIO()

            fig.savefig(
                img_bytes,
                format='png',
                bbox_inches='tight'
            )

            plt.close(fig)

            img_bytes.seek(0)

            img = Image(img_bytes)

            ws.add_image(img, "H2")

    except (
        ValueError,
        TypeError,
        RuntimeError,
        OSError
    ) as e:

        print("Error generando gráfica:", e)

    # Guardar workbook
    wb.save(output)

    output.seek(0)

    return output.getvalue()


# 🔹 EXPORTADORES ESPECÍFICOS

def exportar_excel_biseccion(
    df,
    f_num=None,
    iteraciones=None
):
    return exportar_excel_generico(
        df,
        f_num,
        "Biseccion",
        iteraciones
    )


def exportar_excel_regla_falsa(
    df,
    f_num=None,
    iteraciones=None
):
    return exportar_excel_generico(
        df,
        f_num,
        "ReglaFalsa",
        iteraciones
    )


def exportar_excel_newton(
    df,
    f_num=None,
    iteraciones=None
):
    return exportar_excel_generico(
        df,
        f_num,
        "NewtonRaphson",
        iteraciones
    )


def exportar_excel_secante(
    df,
    f_num=None,
    iteraciones=None
):
    return exportar_excel_generico(
        df,
        f_num,
        "Secante",
        iteraciones
    )


def exportar_excel_punto_fijo(
    df,
    f_num=None,
    gs=None,
    x_min=None,
    x_max=None,
    iteraciones=None
):
    return exportar_excel_generico(
        df,
        f_num,
        "PuntoFijo",
        iteraciones,
        {
            "gs": gs,
            "x_min": x_min,
            "x_max": x_max
        }
    )


def exportar_excel_taylor(
    df,
    f_num=None,
    poly_func=None,
    x_eval=None,
    a=None
):
    return exportar_excel_generico(
        df,
        f_num,
        "Taylor",
        None,
        {
            "poly_func": poly_func,
            "x_eval": x_eval,
            "a": a
        }
    )


def exportar_excel_muller(
    df,
    f_num=None,
    iteraciones=None
):
    return exportar_excel_generico(
        df,
        f_num,
        "Muller",
        iteraciones
    )